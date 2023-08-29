import os
import xml.etree.ElementTree as ET
import openpyxl
from datetime import datetime

# Путь к папке с XML-файлами
folder_path = r"G:\BIM\WIP\03_CIVIL3D\02_Библиотеки\01_Каталоги\02_Безнапорный\ASML_Каталог элементов\Metric Pipes"

# Путь к файлу Excel, который нужно создать
excel_file = r"G:\BIM\WIP\03_CIVIL3D\02_Библиотеки\01_Каталоги\02_Безнапорный\ASML_Каталог элементов\Metric Pipes.xlsx"

# Создание новой книги Excel
wb = openpyxl.Workbook()

# Проверка существования файла Excel и очистка данных при повторном запуске скрипта
if os.path.exists(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # Запись даты предыдущего обновления в ячейку E1
    ws.cell(row=1, column=5).value = f"Предыдущее обновление: {ws.cell(row=2, column=5).value}"
    
    # Очистка данных в листе
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
else:
    ws = wb.active
    ws.title = "Metric Pipes"

# Заполнение заголовков столбцов
ws.cell(row=1, column=1).value = "desc"
ws.cell(row=1, column=2).value = "name"
ws.cell(row=1, column=3).value = "context"
ws.cell(row=1, column=4).value = "XML file"

# Запись даты текущего обновления в ячейку E2
ws.cell(row=2, column=5).value = f"Файл обновлен: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

row = 2

# Список исключений для атрибута desc
exclusions = ["Part Table", "Parameter-driven Display", "Primary Key"]

# Обход всех папок и файлов в указанной директории
for root, dirs, files in os.walk(folder_path):
    for file in files:
        if file.endswith(".xml"):
            file_path = os.path.join(root, file)
            tree = ET.parse(file_path)
            xml_root = tree.getroot()
            
            # Удаление расширения .xml из имени файла
            file_name = os.path.splitext(file)[0]
            
            # Поиск всех элементов в XML-файле
            for elem in xml_root.iter():
                # Поиск атрибутов desc, name и context в элементе
                desc = elem.get("desc")
                name = elem.get("name")
                context = elem.get("context")
                
                # Проверка наличия атрибута desc в списке исключений
                if desc and desc in exclusions:
                    continue
                
                # Запись данных в файл Excel
                if desc or name or context:
                    ws.cell(row=row, column=1).value = desc
                    ws.cell(row=row, column=2).value = name
                    ws.cell(row=row, column=3).value = context
                    ws.cell(row=row, column=4).value = file_name
                    
                    row += 1

# Сохранение файла Excel
wb.save(excel_file)
